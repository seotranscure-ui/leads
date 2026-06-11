"""
Append per-specialty sheets to the master XLSX.

USAGE:
    Open a terminal in the folder containing this script and the XLSX, then:
        pip install openpyxl
        python add_specialty_sheets.py

The script reads "Refined Topical Map - TransCure (Master).xlsx" in the same folder,
filters the master Topical Map sheet for each priority specialty (12 P1 + 15 P2 = 27),
and appends one new sheet per specialty. The original 4 sheets are left untouched.
"""
import os, sys

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Missing dependency. Install with:  pip install openpyxl")
    sys.exit(1)

# ============================================================================
# Constants — priority specialties + the CPT codes commonly billed in each
# ============================================================================

P1_SPECIALTIES = {
    'Pain Management': {'slug':'pain-management','svc':'pain-management',
        'icd':[('M54.5','low back pain'),('M79.7','fibromyalgia'),('G89.21','chronic post-traumatic pain'),
               ('M54.16','lumbar radiculopathy'),('M75.10','rotator cuff tear unspecified')]},
    'Orthopedics': {'slug':'orthopedic','svc':'orthopedic',
        'icd':[('M17.11','primary osteoarthritis right knee'),('M16.11','primary osteoarthritis right hip'),
               ('M25.561','pain in right knee'),('S82.001A','fracture upper end right tibia'),
               ('M51.26','disc displacement lumbar')]},
    'ASC': {'slug':'asc','svc':'surgery-centers/ambulatory',
        'icd':[('Z51.0','radiotherapy session'),('K35.80','acute appendicitis')]},
    'Radiology': {'slug':'radiology','svc':'radiology',
        'icd':[('R10.9','abdominal pain unspecified'),('R07.9','chest pain unspecified')]},
    'Oncology': {'slug':'oncology','svc':'oncology',
        'icd':[('C50.911','breast cancer female right'),('C61','prostate cancer'),
               ('C18.9','colon cancer unspecified'),('C34.90','lung cancer unspecified')]},
    'Gastroenterology': {'slug':'gastroenterology','svc':'gastroenterology',
        'icd':[('K57.30','diverticulosis large intestine'),('K21.9','GERD without esophagitis'),
               ('K50.90','crohn disease unspecified'),('K58.0','IBS with diarrhea')]},
    'Podiatry': {'slug':'podiatry','svc':'podiatry',
        'icd':[('M20.10','hallux valgus'),('L60.0','ingrown nail'),
               ('E11.51','diabetes with peripheral angiopathy'),('M25.572','pain in left ankle')]},
    'Dental': {'slug':'dental','svc':'dental',
        'icd':[('K02.9','dental caries unspecified'),('K05.10','chronic gingivitis')]},
    'Pediatrics': {'slug':'pediatric','svc':'pediatrics',
        'icd':[('Z00.129','encounter routine child health exam'),('J06.9','URI unspecified'),
               ('R50.9','fever unspecified'),('H66.90','otitis media unspecified')]},
    'Cardiology': {'slug':'cardiology','svc':'cardiology',
        'icd':[('I48.91','unspecified atrial fibrillation'),('I50.9','heart failure unspecified'),
               ('I25.10','atherosclerotic heart disease without angina'),('I10','essential hypertension')]},
    'OB/GYN': {'slug':'obgyn','svc':'obgyn',
        'icd':[('Z34.90','encounter supervision normal pregnancy'),('N92.0','excessive menstruation'),
               ('O80','encounter full-term uncomplicated delivery'),('N76.0','acute vaginitis')]},
    'Nephrology': {'slug':'nephrology','svc':'nephrology',
        'icd':[('N18.6','ESRD'),('N40.1','BPH with LUTS'),('N39.0','UTI'),('I12.9','hypertensive CKD')]},
}

P2_SPECIALTIES = {
    'Neurology':'neurology', 'Neurosurgery':'neurosurgery', 'Psychiatry':'psychiatry',
    'Pulmonology':'pulmonology', 'Internal Medicine':'internal-medicine',
    'General Surgery':'general-surgery', 'Primary Care':'primary-care',
    'Dermatology':'dermatology', 'Ophthalmology':'ophthalmology',
    'Endocrinology':'endocrinology', 'Anesthesia':'anesthesia',
    'Family Practice':'family-practice', 'Urgent Care':'urgent-care',
    'Infectious Disease':'infectious-disease', 'Mental Health':'mental-health',
}

# ============================================================================
# Configuration
# ============================================================================
HERE = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(HERE, 'Refined Topical Map - TransCure (Master).xlsx')

# ============================================================================
# Styles
# ============================================================================
H_FILL = PatternFill('solid', fgColor='1F4E78')
H_FONT = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
SECTION_FILL = PatternFill('solid', fgColor='FFE699')
WRAP = Alignment(wrap_text=True, vertical='top')
WAVE_FILLS = {1:PatternFill('solid',fgColor='F4B084'),2:PatternFill('solid',fgColor='FFE699'),
              3:PatternFill('solid',fgColor='C6EFCE'),4:PatternFill('solid',fgColor='BDD7EE')}
PRI_FILLS = {'P0':PatternFill('solid',fgColor='C00000'),'P1':PatternFill('solid',fgColor='F4B084'),
             'P2':PatternFill('solid',fgColor='FFE699'),'P3':PatternFill('solid',fgColor='E2EFDA')}
STATUS_FILLS = {'LIVE':PatternFill('solid',fgColor='C6EFCE'),'NEW':PatternFill('solid',fgColor='FFEB9C')}

HEADERS = [
    'Wave','Topics','Parent URL Slug','Slug','Title (≤60 ch)','Meta Description (≤130 ch)',
    'Image URL','Image Alt','Image Concept',
    'Central Entity','Entity Type','Canonical Query','Query Template Class','Intent',
    'Core H2s (Question Form)',
    'S-node links','C-node links','I-node links','Incoming links',
    'Live Status','Priority','Cluster'
]
COL_WIDTHS = [6,28,32,40,48,48,36,32,38,26,22,26,22,20,48,28,42,42,52,10,8,28]

# Section order (cluster keys to bucket into headings)
SECTION_ORDER = [
    ('SERVICE — NATIONAL', ['Core/Services-Specialty']),
    ('SERVICE — TIER-1 STATES', ['Core/Services-Specialty-State']),
    ('SERVICE — CITIES', ['Core/Services-Specialty-City']),
    ('INFORMATIONAL — HUB', ['Outer/Specialty-Hub']),
    ('INFORMATIONAL — CPT CHEAT SHEET', ['Outer/Specialty-CPT-Cheatsheet']),
    ('INFORMATIONAL — MODIFIERS', ['Outer/Specialty-Modifier']),
    ('INFORMATIONAL — MISTAKES', ['Outer/Specialty-Mistakes']),
    ('INFORMATIONAL — KPIs', ['Outer/Specialty-KPIs']),
    ('INFORMATIONAL — COMMON DENIALS', ['Outer/Specialty-Denials']),
    ('INFORMATIONAL — PRIOR AUTHORIZATION', ['Outer/Specialty-PriorAuth']),
    ('INFORMATIONAL — CREDENTIALING', ['Outer/Specialty-Credentialing']),
    ('INFORMATIONAL — AUDIT', ['Outer/Specialty-Audit']),
    ('INFORMATIONAL — RCM', ['Outer/Specialty-RCM']),
    ('INFORMATIONAL — MEDICARE', ['Outer/Specialty-Medicare']),
    ('COMPARISON — INHOUSE VS OUTSOURCE', ['Outer/Specialty-Compare']),
    ('COMPARATIVE — COMPANIES LISTICLE', ['Outer/Companies-Specialty-Listicle']),
]

# ============================================================================
# Load master sheet → build a list of topic dicts
# ============================================================================
def load_master_topics(wb):
    ws = wb['1. Topical Map (Master)']
    headers = [ws.cell(1, i).value for i in range(1, ws.max_column+1)]
    topics = []
    for r in range(2, ws.max_row+1):
        if ws.cell(r, 4).value is None:
            continue
        row = {}
        for i, h in enumerate(headers, 1):
            row[h] = ws.cell(r, i).value
        topics.append(row)
    return topics

# ============================================================================
# Filter helpers
# ============================================================================
def is_specialty_topic(t, sl, svc):
    s = t.get('Slug') or ''
    cluster = t.get('Cluster') or ''
    # Service pages for this specialty
    if cluster.startswith('Core/Services-Specialty'):
        if s == f'/medical-billing/services/{svc}/' or s.startswith(f'/medical-billing/services/{svc}/'):
            return True
    # Outer specialty pages
    if cluster.startswith('Outer/Specialty-') or cluster == 'Outer/Companies-Specialty-Listicle':
        if s == f'/medical-billing/{sl}/' or s.startswith(f'/medical-billing/{sl}-') or s == f'/medical-billing/companies-{sl}/':
            return True
    return False

def cpt_pages_for(topics, spec_slug, spec_cpts):
    if not spec_cpts:
        return []
    by_slug = {t.get('Slug'): t for t in topics}
    pages = []
    for code in spec_cpts:
        s = f'/medical-billing/code/cpt/{code}/'
        if s in by_slug:
            pages.append(by_slug[s])
    return pages

def icd_pages_for(topics, icds):
    by_slug = {t.get('Slug'): t for t in topics}
    pages = []
    for icd, cond in icds:
        s = f'/medical-billing/code/icd-10/{icd.lower().replace(".","-")}/'
        if s in by_slug:
            pages.append(by_slug[s])
    return pages

# Specialty -> CPT codes commonly billed (top codes used in topic generation)
SPECIALTY_CPT_CODES = {
    'cardiology': ['93306','93000','93005','93010','93015','93016','93017','93018','93350','93454',
                    '93458','93656','33208','33249','33533','92928','92960','93320','93325','93798'],
    'orthopedic': ['27130','27447','29881','29888','22612','23472','20610','20605','20680','63030',
                   '63047','27650','25000','29515','15847'],
    'pain-management': ['62321','64490','64520','64640','64721','64999','20550','20605','20610',
                        '95819','95816','95886','97163','97164'],
    'oncology': ['96365','96366','96367','96372','96375','96402','96413','96415','77386','77300',
                 '77065','77067'],
    'radiology': ['70450','70551','70553','71045','71046','71250','71260','72040','72100','72148',
                  '73030','73502','73562','73721','74176','74177','76536','76641','76700','76705',
                  '76801','76856','76942','77063','77067'],
    'gastroenterology': ['43249','45378','45380','45385','43239','43235','45330','45331','47562',
                         '47563','91035','44388','43200'],
    'podiatry': ['11055','11056','11057','11721','11730','11719','11720','11750','28092','97597',
                 '11044','11045','99213','99214'],
    'dental': ['D0120','D0150','D0210','D0220','D0274','D1110','D1120','D1206','D1351','D2140',
               'D2150','D2330','D2391','D7140','D7210','D7240'],
    'pediatric': ['99381','99382','99383','99391','99392','99393','99394','99213','99214','99203',
                  '99204','90471','90472','90460','90461'],
    'obgyn': ['58571','58573','57454','57500','59409','59426','58100','58301','76801','76817',
              '76830','59400','59510','59515','59812'],
    'asc': ['45378','66984','43239','29881','62321','64483','64635','29826','29827'],
    'nephrology': ['90935','90937','90945','90951','90952','90960','90961','90962','50390','99213'],
}

# ============================================================================
# Write one specialty sheet
# ============================================================================
def write_specialty_sheet(wb, spec_name, spec_info, topics, is_p2=False):
    if isinstance(spec_info, dict):
        sl = spec_info['slug']
        svc = spec_info['svc']
        icds = spec_info.get('icd', [])
    else:
        sl = spec_info
        svc = spec_info
        icds = []

    sheet_name = (f'Spec — {spec_name}')[:31]
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    # Title row
    ws.cell(1, 1, f'Topical Map — {spec_name} Medical Billing')
    ws.cell(1, 1).font = Font(bold=True, size=16, color='1F4E78')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=22)

    subtitle = f'Informational hub: /medical-billing/{sl}/   ·   Commercial service: /medical-billing/services/{svc}/'
    ws.cell(2, 1, subtitle)
    ws.cell(2, 1).font = Font(italic=True, color='595959')
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=22)

    HEADER_ROW = 4
    for i, h in enumerate(HEADERS, 1):
        c = ws.cell(HEADER_ROW, i, h)
        c.font = H_FONT
        c.fill = H_FILL
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[HEADER_ROW].height = 36
    ws.freeze_panes = f'A{HEADER_ROW+1}'

    r = HEADER_ROW + 1
    total = 0

    def write_row(t):
        nonlocal r, total
        vals = [
            t.get('Wave'), t.get('Topics'), t.get('Parent URL Slug'), t.get('Slug'),
            t.get('Title (≤60 ch)'), t.get('Meta Description (≤130 ch)'),
            t.get('Image URL'), t.get('Image Alt'), t.get('Image Concept'),
            t.get('Central Entity'), t.get('Entity Type'), t.get('Canonical Query'),
            t.get('Query Template Class'), t.get('Intent'), t.get('Core H2s (Question Form)'),
            t.get('S-node links'), t.get('C-node links'), t.get('I-node links'),
            t.get('Incoming links'), t.get('Live Status'), t.get('Priority'), t.get('Cluster'),
        ]
        for i, v in enumerate(vals, 1):
            ws.cell(r, i, v)
        wave = t.get('Wave')
        if wave in WAVE_FILLS: ws.cell(r, 1).fill = WAVE_FILLS[wave]
        pri = t.get('Priority')
        if pri in PRI_FILLS: ws.cell(r, 21).fill = PRI_FILLS[pri]
        st = t.get('Live Status')
        if st in STATUS_FILLS: ws.cell(r, 20).fill = STATUS_FILLS[st]
        for col in [2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 22]:
            ws.cell(r, col).alignment = WRAP
        r += 1
        total += 1

    def write_section_header(label):
        nonlocal r
        ws.cell(r, 1, label)
        ws.cell(r, 1).font = Font(bold=True, size=11, color='1F4E78')
        ws.cell(r, 1).fill = SECTION_FILL
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=22)
        r += 1

    # Filter master topics for this specialty
    spec_topics = [t for t in topics if is_specialty_topic(t, sl, svc)]
    by_cluster = {}
    for t in spec_topics:
        by_cluster.setdefault(t.get('Cluster',''), []).append(t)

    for label, cluster_keys in SECTION_ORDER:
        section_rows = []
        for ck in cluster_keys:
            section_rows.extend(by_cluster.get(ck, []))
        if section_rows:
            write_section_header(label)
            for t in sorted(section_rows, key=lambda x: x.get('Slug') or ''):
                write_row(t)

    # Cross-referenced individual CPT codes for this specialty
    cpts = cpt_pages_for(topics, sl, SPECIALTY_CPT_CODES.get(sl, []))
    if cpts:
        write_section_header(f'CODE LIBRARY — Individual CPT codes commonly billed in {spec_name} ({len(cpts)})')
        for t in sorted(cpts, key=lambda x: x.get('Slug') or ''):
            write_row(t)

    # Cross-referenced ICD-10 condition pages
    icd_pgs = icd_pages_for(topics, icds)
    if icd_pgs:
        write_section_header(f'CODE LIBRARY — ICD-10 conditions for {spec_name} ({len(icd_pgs)})')
        for t in sorted(icd_pgs, key=lambda x: x.get('Slug') or ''):
            write_row(t)

    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    if r > HEADER_ROW + 1:
        ws.auto_filter.ref = f'A{HEADER_ROW}:V{r-1}'
    return total

# ============================================================================
# Main
# ============================================================================
def main():
    if not os.path.exists(XLSX_PATH):
        print(f"ERROR: Cannot find {XLSX_PATH}")
        print(f"Place this script in the same folder as the master XLSX.")
        sys.exit(1)
    print(f'Loading {XLSX_PATH}...')
    wb = load_workbook(XLSX_PATH)
    print(f'Found {len(wb.sheetnames)} existing sheets: {wb.sheetnames}')
    topics = load_master_topics(wb)
    print(f'Loaded {len(topics)} topics from master sheet')

    # Write P1 sheets
    print('\n=== P1 specialty sheets ===')
    for spec_name, info in P1_SPECIALTIES.items():
        n = write_specialty_sheet(wb, spec_name, info, topics, is_p2=False)
        print(f'  {spec_name}: {n} topics')

    # Write P2 sheets
    print('\n=== P2 specialty sheets ===')
    for spec_name, slug in P2_SPECIALTIES.items():
        n = write_specialty_sheet(wb, spec_name, {'slug':slug, 'svc':slug, 'icd':[]}, topics, is_p2=True)
        print(f'  {spec_name}: {n} topics')

    wb.save(XLSX_PATH)
    print(f'\nSaved {XLSX_PATH}')
    print(f'Total sheets: {len(wb.sheetnames)}')
    print(f'File size: {os.path.getsize(XLSX_PATH)/1024:.1f} KB')

if __name__ == '__main__':
    main()
